from django.shortcuts import render, get_object_or_404
from Costeo.models import Depreciacion, FactoresPremisas, CostosOperativos, Costeo
from Costeo.forms import CosteoForm, CosteoActualizarForm
from django.contrib import messages
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from Rutas.models import Rutas, TiposUnidades, Modelos
from django.template.loader import render_to_string
from urllib.request import urlopen
import xmltodict
import json
import requests
from decimal import Decimal
from datetime import datetime
from django.conf import settings


import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.cell import Cell
from openpyxl.styles import Alignment

from django.utils.encoding import smart_str

# Create your views here.
def Home(request):
    if (request.user.is_authenticated):
        return render(request, "costeoHome.html")
    else:
	    return HttpResponseRedirect('/')

def tablaCostos(request):
    if (request.user.is_authenticated):
        data = dict()
        costeos = Costeo.objects.filter(Actividad = 1).order_by('Empresa').select_related('IDDepreciacion').select_related('IDRuta').select_related('IDFactoresPremisas').select_related('IDCostosOperativos').all()
        data['tablaCostos'] = render_to_string('tablaCostos.html', {'costeos': costeos})
        return JsonResponse(data)
    else:
	    return HttpResponseRedirect('/')

def excelGeneral(request):
    if (request.user.is_authenticated):
        costeos = Costeo.objects.filter(Actividad=1).select_related('IDDepreciacion').select_related('IDRuta').select_related('IDFactoresPremisas').select_related('IDCostosOperativos').all()
        libro = Workbook()
        excel_document = openpyxl.load_workbook(settings.EXCELGENERAL)
        libro = excel_document
        puntero = libro.active

        cont = 2 
        for costeo in costeos:
            puntero['A'+str(cont)] = costeo.Empresa
            puntero['B'+str(cont)] = costeo.IDRuta.NombreRuta
            puntero['C'+str(cont)] = costeo.IDTipoUnidad.Nombre
            puntero['D'+str(cont)] = costeo.Kilometros
            puntero['E'+str(cont)] = costeo.TotalCostos
            puntero['F'+str(cont)] = costeo.FactorAjustePor
            puntero['G'+str(cont)] = costeo.FactorAjuste
            puntero['H'+str(cont)] = costeo.Mop
            puntero['I'+str(cont)] = costeo.MopPor
            puntero['J'+str(cont)] = costeo.TotalTransportista
            cont = cont + 1

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="excelGeneral-'+ str(datetime.now())+'.xlsx"'
        libro.save(response)
        return response
    else:
	    return HttpResponseRedirect('/')

def excelCosteo(request,id):
    if (request.user.is_authenticated):
        costeos = Costeo.objects.filter(id=id).select_related('IDDepreciacion').select_related('IDRuta').select_related('IDFactoresPremisas').select_related('IDCostosOperativos').all()
        libro = Workbook()
        excel_document = openpyxl.load_workbook(settings.ARCHIVOMAESTRO)
        libro = excel_document
        puntero = libro.active

        puntero['B2'] = costeos[0].IDRuta.NombreRuta
        puntero['B3'] = costeos[0].Kilometros
        if(costeos[0].ViajeRedondo == True):
            puntero['A4'] = "Viaje Redondo"
        else:
            puntero['A4'] = "Viaje Sencillo"
        puntero['A5'] = costeos[0].Producto
        puntero['B5'] = costeos[0].IDTipoUnidad.Nombre

        puntero['B7'] = costeos[0].IDDepreciacion.Meses
        puntero['B8'] = costeos[0].IDDepreciacion.CostosUnidad
        puntero['B9'] = costeos[0].IDDepreciacion.CostosCaja
        puntero['B10'] = costeos[0].IDDepreciacion.ViajesMes
        puntero['B11'] = costeos[0].IDDepreciacion.KmsMesXunidad


        puntero['B17'] = costeos[0].IDDepreciacion.RentaGPS
        puntero['B18'] = costeos[0].IDDepreciacion.PlacasTenencia
        puntero['B19'] = costeos[0].IDDepreciacion.Seguro
        puntero['B20'] = costeos[0].IDDepreciacion.Admvo/100
        puntero['B21'] = costeos[0].IDDepreciacion.Financieros/100
        puntero['B22'] = costeos[0].IDDepreciacion.MttoUnidadXkm
        puntero['B23'] = costeos[0].IDDepreciacion.Llantas
        puntero['B24'] = costeos[0].IDDepreciacion.Operador

        puntero['B29'] = costeos[0].IDFactoresPremisas.Km
        puntero['B32'] = costeos[0].IDFactoresPremisas.CasetaSingle
        puntero['B33'] = costeos[0].IDFactoresPremisas.Rendimiento
        puntero['B34'] = costeos[0].IDFactoresPremisas.Diesel 

        puntero['A59'] = costeos[0].MopPor/100
        puntero['A60'] = costeos[0].FactorAjustePor/100

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="excel-'+str(costeos[0].Empresa)+'-'+str(datetime.now())+'.xlsx"'
        libro.save(response)
        return response
    else:
	    return HttpResponseRedirect('/')

import pdfkit
def pdfCosteo(request,id):
    if (request.user.is_authenticated):
        opciones = {
            'page-size': 'Letter',
            'margin-top': '0.40in',
            'margin-right': '0.75in',
            'margin-bottom': '0.10in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
        }
        path_wkhtmltopdf = settings.PDFKIT
        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
        pdf = pdfkit.from_url (settings.RUTAPDF+str(id)+'/', False, options = opciones)
        response = HttpResponse(pdf,content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="documento.pdf"'
        return response
    else:
	    return HttpResponseRedirect('/')

def hojaPdfCosteo(request, id):
    costeo = get_object_or_404(Costeo, pk=id)
    return render(request,'hojaPdfCosteo.html', {'costeo': costeo})


#Crear un costeo
def costeos(request):
    if (request.user.is_authenticated):
        depreciacion = get_object_or_404(Depreciacion, pk=1)
        factoresPremisas = get_object_or_404(FactoresPremisas, pk=1)

        file = urlopen('https://publicacionexterna.azurewebsites.net/publicaciones/prices')
        data = file.read()
        file.close()
        MaxDiesel=0
        count=0
        data = xmltodict.parse(data)
        for Combustible in data['places']['place'] :
            for precios in Combustible['gas_price']:
                if isinstance(precios, str) is False:
                    if (precios['@type']) == 'diesel':
                        PrecioMax=float(precios['#text'])
                        MaxDiesel+=PrecioMax
                        count+=1

        
        costeo = Costeo.objects.all()
        MaxDiesel=MaxDiesel/count
        MaxDiesel=round(MaxDiesel,2)
        factoresPremisas.Diesel=MaxDiesel
        factoresPremisas.DieselSinIva=round(MaxDiesel*(1-.16),2)
        form = CosteoForm()

        if request.method == "POST":
            form = CosteoForm(request.POST)
            if form.is_valid():
                depreciacion = Depreciacion(
                    Meses=request.POST['mesesId'],
                    CostosUnidad=request.POST['costosUnidadId'],
                    CostosCaja=request.POST['costosCajaId'],
                    ViajesMes=request.POST['viajesMesId'],
                    KmsMesXunidad=request.POST['kmsMesXunidadId'],
                    KmsMaximo=request.POST['kmsMaximoId'],
                    DepTracto=request.POST['depTractoId'],
                    DepCaja=request.POST['depCajaId'],
                    RentaGPS=request.POST['rentaGPSId'],
                    PlacasTenencia=request.POST['placasTenenciaId'],
                    Seguro=request.POST['seguroId'],
                    Admvo=request.POST['admvoId'],
                    Financieros=request.POST['financierosId'],
                    MttoUnidadXkm=request.POST['mttoUnidadXkmId'],
                    Llantas=request.POST['llantasId'],
                    Operador=request.POST['operadorId'],
                    IDUsuarioAlta = request.user.id,
                    FechaAlta = datetime.now(),
                )
                depreciacion.save()

                factoresPremisas = FactoresPremisas(
                    Unidad=request.POST['unidadId'],
                    Caja=1,
                    Km=request.POST['kmSencilloId'],
                    KmMensuales=request.POST['kmMensualesId'],
                    CasetaSingle=request.POST['casetaSingleId'],
                    Rendimiento=request.POST['rendimientoId'],
                    Diesel=request.POST['dieselId'],
                    DieselSinIva=request.POST['dieselSinIvaId'],
                    IDUsuarioAlta = request.user.id,
                    FechaAlta = datetime.now(),
                )
                factoresPremisas.save()
                costosOperativos = CostosOperativos(
                    Combustible=request.POST['CombustibleId'],
                    Casetas=request.POST['CasetasId'],
                    Operador=request.POST['OperadorId'],
                    Subtotal1=request.POST['SubTotalId'],
                    MttoUnidad=request.POST['MttoUnidadId'],
                    LlantasUnidad=request.POST['LlantasUnidadId'],
                    Gps=request.POST['GPSId'],
                    Seguro=request.POST['SeguroId'],
                    PlacasTenencia=request.POST['PlacasTenenciaCostId'],
                    SubTotal2=request.POST['SubTotal2Id'],
                    Admvo=request.POST['AdmvoId'],
                    Financieros=request.POST['FinancierosId'],
                    DeprUnidad=request.POST['DeprUnidadId'],
                    DeprRemolque=request.POST['DeprRemolqueId'],
                    Subtotal3=request.POST['SubTotal3Id'],
                    IDUsuarioAlta = request.user.id,
                    FechaAlta = datetime.now(),
                )
                costosOperativos.save()

                if(form.cleaned_data["IDModelo"] == 0):
                    costeo = Costeo(
                        IDRuta=get_object_or_404(Rutas, pk=form.cleaned_data["IDRuta"]),
                        CPOrigen = form.cleaned_data["CPOrigen"],
                        CPDestino = form.cleaned_data["CPDestino"],
                        Empresa = form.cleaned_data["Empresa"],
                        Casetas = form.cleaned_data["Casetas"],
                        IDDepreciacion=depreciacion,
                        IDFactoresPremisas=factoresPremisas,
                        IDCostosOperativos=costosOperativos,
                        IDTipoUnidad=get_object_or_404(TiposUnidades, pk=form.cleaned_data["IDTipoUnidad"]),
                        Kilometros = form.cleaned_data["Kilometros"],
                        Producto=form.cleaned_data["Producto"],
                        TotalCostos=form.cleaned_data["TotalCostos"],
                        FactorAjustePor=form.cleaned_data["FactorAjustePor"],
                        FactorAjuste=form.cleaned_data["FactorAjuste"],
                        Mop=form.cleaned_data["Mop"],
                        MopPor=form.cleaned_data["MopPor"],
                        TotalTransportista=form.cleaned_data["TotalTransportista"],
                        ViajeRedondo = form.cleaned_data["ViajeRedondo"],
                        IDUsuarioAlta = request.user.id,
                        FechaAlta = datetime.now(),
                    )
                    costeo.save()
                else:
                    costeo = Costeo(
                        IDRuta=get_object_or_404(Rutas, pk=form.cleaned_data["IDRuta"]),
                        CPOrigen = form.cleaned_data["CPOrigen"],
                        CPDestino = form.cleaned_data["CPDestino"],
                        Empresa = form.cleaned_data["Empresa"],
                        Casetas = form.cleaned_data["Casetas"],
                        IDDepreciacion=depreciacion,
                        IDFactoresPremisas=factoresPremisas,
                        IDCostosOperativos=costosOperativos,
                        IDTipoUnidad=get_object_or_404(TiposUnidades, pk=form.cleaned_data["IDTipoUnidad"]),
                        IDModelo=get_object_or_404(Modelos, pk=form.cleaned_data["IDModelo"]),
                        Kilometros = form.cleaned_data["Kilometros"],
                        Producto=form.cleaned_data["Producto"],
                        TotalCostos=form.cleaned_data["TotalCostos"],
                        FactorAjustePor=form.cleaned_data["FactorAjustePor"],
                        FactorAjuste=form.cleaned_data["FactorAjuste"],
                        Mop=form.cleaned_data["Mop"],
                        MopPor=form.cleaned_data["MopPor"],
                        TotalTransportista=form.cleaned_data["TotalTransportista"],
                        ViajeRedondo = form.cleaned_data["ViajeRedondo"],
                        IDUsuarioAlta = request.user.id,
                        FechaAlta = datetime.now(),
                    )
                    costeo.save()
                return HttpResponseRedirect('/costeo/Home')
            else:
                print(messages.error(request, "Error"))

        context = {"form": form, "costeo": costeo, 'depreciacion': depreciacion, 'FactoresPremisas': factoresPremisas}
        return render(request, "Costeo.html", context)
    else:
        return HttpResponseRedirect('/')

def selectRutasCosto(request):
    if (request.user.is_authenticated):
        data = dict()
        rutas = Rutas.objects.filter(Actividad = 1).order_by('NombreRuta').all()
        context = {"rutas": rutas}
        data['selectRutasCosto'] = render_to_string('selectRutasCosto.html',context, request=request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def buscarRutaCosto(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        ruta = get_object_or_404(Rutas, pk=datos['idR'])
        data['CPOrigen'] = ruta.CPOrigen
        data['EstadoOrigen'] = ruta.EstadoOrigen
        data['CiudadOrigen'] = ruta.CiudadOrigen
        data['CPDestino'] = ruta.CPDestino
        data['EstadoDestino'] = ruta.EstadoDestino
        data['CiudadDestino'] = ruta.CiudadDestino
        data['Kilometros'] = ruta.Kilometros
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def selectTiposUnidadesCosto(request):
    if (request.user.is_authenticated):
        data = dict()
        tiposUnidades = TiposUnidades.objects.filter(Actividad = 1).order_by('Nombre').all()
        context = {'tiposUnidades': tiposUnidades}
        data['selectTiposUnidadesCosto'] = render_to_string('selectTiposUnidadesCosto.html',context, request=request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def selectModelosCosteo(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        tiposUnidades = TiposUnidades.objects.filter(id = datos['IDTipoUnidad'], Actividad = 1).order_by('Nombre').prefetch_related('IDModelo').all()
        context = {'tiposUnidades': tiposUnidades[0]}
        data['selectModelosCosteo'] = render_to_string('selectModelosCosteo.html',context, request=request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')


#Rutas Filtro En Costo
def selectRutasFiltroCostos(request):
    if (request.user.is_authenticated):
        data = dict()
        rutas = Rutas.objects.filter(Actividad = 1).order_by('NombreRuta').distinct()
        data['selectRutasFiltroCostos'] = render_to_string('selectRutasFiltroCostos.html', {'rutas': rutas})
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def selectTipoUnidadesFiltroCostos(request):
    if (request.user.is_authenticated):
        tiposUnidades = TiposUnidades.objects.filter(Actividad = 1).order_by('Nombre').all()
        data = dict()
        data['selectTipoUnidadesFiltroCostos'] = render_to_string('selectTipoUnidadesFiltroCostos.html', {'tiposUnidades': tiposUnidades})
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def filtroCortoCostos(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        costeos = Costeo.objects.filter(IDRuta=datos['IDRuta'], Actividad =1, IDTipoUnidad=datos['IDTipoUnidad'], TotalTransportista__range=(Decimal(datos['LimiteInferior']), Decimal(datos['LimiteSuperior']))).all()
        data['filtroCortoCostos'] = render_to_string('filtroLCCosteo.html', {'costeos': costeos})
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def filtroLargoCostos(request):
    if (request.user.is_authenticated):
        request_data = request.POST
        datos = json.loads(request_data['data'])
        data = dict()
        rutas = Rutas.objects.filter(Actividad = 1, CPOrigen = datos['CPOrigen'], EstadoOrigen=datos['EstadoOrigen'], CiudadOrigen=datos['CiudadOrigen'], CPDestino = datos['CPDestino'], EstadoDestino=datos['EstadoDestino'], CiudadDestino=datos['CiudadDestino']).all()
        costeos = Costeo.objects.filter(IDRuta= rutas[0], Actividad = 1, IDTipoUnidad=datos['IDTipoUnidad'], TotalTransportista__range=(Decimal(datos['LimiteInferior']), Decimal(datos['LimiteSuperior'])))
        data['filtroLargoCostos'] = render_to_string('filtroLCCosteo.html', {'costeos': costeos})
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')



#Actualizar costeo
def updateCosteo(request, id):
    if (request.user.is_authenticated):
        costeo=get_object_or_404(Costeo, pk=id)
        depreciacion = get_object_or_404(Depreciacion, pk=costeo.IDDepreciacion.id)
        factoresPremisas = get_object_or_404(FactoresPremisas, pk=costeo.IDFactoresPremisas.id)
        costosOperativos = get_object_or_404(CostosOperativos, pk=costeo.IDCostosOperativos.id)
        form = CosteoActualizarForm(instance=costeo)

        if request.method == "POST":
            form = CosteoActualizarForm(request.POST)
            if form.is_valid():
                
                depreciacion.Meses=request.POST['mesesId']
                depreciacion.CostosUnidad=request.POST['costosUnidadId']
                depreciacion.CostosCaja=request.POST['costosCajaId']
                depreciacion.ViajesMes=request.POST['viajesMesId']
                depreciacion.KmsMesXunidad=request.POST['kmsMesXunidadId']
                depreciacion.KmsMaximo=request.POST['kmsMaximoId']
                depreciacion.DepTracto=request.POST['depTractoId']
                depreciacion.DepCaja=request.POST['depCajaId']
                depreciacion.RentaGPS=request.POST['rentaGPSId']
                depreciacion.PlacasTenencia=request.POST['placasTenenciaId']
                depreciacion.Seguro=request.POST['seguroId']
                depreciacion.Admvo=request.POST['admvoId']
                depreciacion.Financieros=request.POST['financierosId']
                depreciacion.MttoUnidadXkm=request.POST['mttoUnidadXkmId']
                depreciacion.Llantas=request.POST['llantasId']
                depreciacion.Operador=request.POST['operadorId']
                depreciacion.IDUsuarioMod = request.user.id
                depreciacion.FechaModificacion = datetime.now()
                depreciacion.save()

                factoresPremisas.Unidad=request.POST['unidadId']
                factoresPremisas.Km=request.POST['kmSencilloId']
                factoresPremisas.KmRoundTrip=request.POST['kmRoundTripId']
                factoresPremisas.KmMensuales=request.POST['kmMensualesId']
                factoresPremisas.CasetaSingle=request.POST['casetaSingleId']
                factoresPremisas.Rendimiento=request.POST['rendimientoId']
                factoresPremisas.Diesel=request.POST['dieselId']
                factoresPremisas.DieselSinIva=request.POST['dieselSinIvaId']
                factoresPremisas.IDUsuarioMod = request.user.id
                factoresPremisas.FechaModificacion = datetime.now()
                factoresPremisas.save()

                costosOperativos.Combustible=request.POST['CombustibleId']
                costosOperativos.Casetas=request.POST['CasetasId']
                costosOperativos.Operador=request.POST['OperadorId']
                costosOperativos.Subtotal1=request.POST['SubTotalId']
                costosOperativos.MttoUnidad=request.POST['MttoUnidadId']
                costosOperativos.LlantasUnidad=request.POST['LlantasUnidadId']
                costosOperativos.Gps=request.POST['GPSId']
                costosOperativos.Seguro=request.POST['SeguroId']
                costosOperativos.PlacasTenencia=request.POST['PlacasTenenciaCostId']
                costosOperativos.SubTotal2=request.POST['SubTotal2Id']
                costosOperativos.Admvo=request.POST['AdmvoId']
                costosOperativos.Financieros=request.POST['FinancierosId']
                costosOperativos.DeprUnidad=request.POST['DeprUnidadId']
                costosOperativos.DeprRemolque=request.POST['DeprRemolqueId']
                costosOperativos.Subtotal3=request.POST['SubTotal3Id']
                costosOperativos.IDUsuarioMod = request.user.id
                costosOperativos.FechaModificacion = datetime.now()
                costosOperativos.save()

                costeo.IDRuta = get_object_or_404(Rutas, pk= form.cleaned_data["IDRuta"])
                costeo.CPOrigen = form.cleaned_data["CPOrigen"]
                costeo.CPDestino = form.cleaned_data["CPDestino"]
                costeo.Empresa = form.cleaned_data["Empresa"]
                costeo.Casetas = form.cleaned_data["Casetas"]
                costeo.IDTipoUnidad = get_object_or_404(TiposUnidades, pk= form.cleaned_data["IDTipoUnidad"])
                if form.cleaned_data["IDModelo"] == 0:
                    if costeo.IDModelo!=None:
                        costeo.IDModelo = None
                else:
                    costeo.IDModelo = get_object_or_404(Modelos, pk= form.cleaned_data["IDModelo"])
                costeo.Kilometros=form.cleaned_data["Kilometros"]
                costeo.Producto=form.cleaned_data["Producto"]
                costeo.TotalCostos=form.cleaned_data["TotalCostos"]
                costeo.FactorAjustePor=form.cleaned_data["FactorAjustePor"]
                costeo.FactorAjuste=form.cleaned_data["FactorAjuste"] 
                costeo.Mop=form.cleaned_data["Mop"]
                costeo.MopPor=form.cleaned_data["MopPor"] 
                costeo.TotalTransportista=form.cleaned_data["TotalTransportista"]
                costeo.ViajeRedondo = form.cleaned_data["ViajeRedondo"]
                costeo.IDUsuarioMod = request.user.id
                costeo.FechaModificacion = datetime.now()
                costeo.save()
                return HttpResponseRedirect('/costeo/Home')
            else:
                print(messages.error(request, "Error"))
            
        context = {"form": form, "costeo": costeo, 'depreciacion': depreciacion, 'FactoresPremisas': factoresPremisas}
        return render(request, "Costeo2.html", context)
    else:
        return HttpResponseRedirect('/')

def selectRutasCostoActualizar(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        rutas = Rutas.objects.filter(Actividad = 1).order_by('NombreRuta').all()
        context = {"rutas": rutas, "idR" : int(datos['idR'])}
        data['selectRutasCostoActualizar'] = render_to_string('selectRutasCostoActualizar.html',context, request=request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def buscarRutaCostoActualizar(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        ruta = get_object_or_404(Rutas, pk=datos['idR'])
        data['CPOrigen'] = ruta.CPOrigen
        data['EstadoOrigen'] = ruta.EstadoOrigen
        data['CiudadOrigen'] = ruta.CiudadOrigen
        data['CPDestino'] = ruta.CPDestino
        data['EstadoDestino'] = ruta.EstadoDestino
        data['CiudadDestino'] = ruta.CiudadDestino
        data['Kilometros'] = ruta.Kilometros
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def selectTiposUnidadesCostoActualizar(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        tiposUnidades = TiposUnidades.objects.filter(Actividad = 1).order_by('Nombre').all()
        context = {'tiposUnidades': tiposUnidades, 'idTU' : int(datos['idTU'])}
        data['selectTiposUnidadesCostoActualizar'] = render_to_string('selectTiposUnidadesCostoActualizar.html',context, request=request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/')

def selectModelosCosteoActualizar(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])
        tiposUnidades = TiposUnidades.objects.filter(id = datos['IDTipoUnidad'], Actividad = 1).order_by('Nombre').prefetch_related('IDModelo').all()
        context = {'tiposUnidades': tiposUnidades[0], 'IDModelo':datos['IDModelo']}
        data['selectModelosCosteoActualizar'] = render_to_string('selectModelosCosteoActualizar.html',context, request=request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/') 

def borrarCosto(request,id):
    if (request.user.is_authenticated):
        costeo = get_object_or_404(Costeo, pk=id)

        depreciacion = get_object_or_404(Depreciacion, pk=costeo.IDDepreciacion.id)
        depreciacion.IDUsuarioMod = request.user.id
        depreciacion.FechaModificacion = datetime.now()
        depreciacion.Actividad = 0
        depreciacion.save()

        factoresPremisas = get_object_or_404(FactoresPremisas, pk=costeo.IDFactoresPremisas.id)
        factoresPremisas.IDUsuarioMod = request.user.id
        factoresPremisas.FechaModificacion = datetime.now()
        factoresPremisas.Actividad = 0
        factoresPremisas.save()

        costosOperativos = get_object_or_404(CostosOperativos, pk=costeo.IDCostosOperativos.id)
        costosOperativos.IDUsuarioMod = request.user.id
        costosOperativos.FechaModificacion = datetime.now()
        costosOperativos.Actividad = 0
        costosOperativos.save()

        costeo.IDUsuarioMod = request.user.id
        costeo.FechaModificacion = datetime.now()
        costeo.Actividad = 0
        costeo.save()

        return HttpResponseRedirect('/costeo/Home')
    else:
        return HttpResponseRedirect('/')   



def IdCdCosto(request):
    if (request.user.is_authenticated):
        data = dict()
        request_data = request.POST
        datos = json.loads(request_data['data'])

        url = 'http://gaia.inegi.org.mx/sakbe_v3.1/buscadestino' 
        params = {'type': "json",'buscar':datos['ciudad'], 'proj': "MERC",'num':50,'key': "x4cGBaOC-423Z-A9ii-AKwM-QWiCfq63rMu9"}
        r = requests.post(url, params=params)
        resjson = r.json()
        for val in resjson['data']:
            string=val['ent_abr'].replace(".","")
            string=string.replace(" ","")
            if comparaestadosCosto(string)==datos['estado']:
                data["cod"] = val['id_dest']
                data["bandera"] = 2
                return JsonResponse(data)

        palabras = datos['ciudad'].split()
        for palabra in palabras:
            url = 'http://gaia.inegi.org.mx/sakbe_v3.1/buscadestino' 
            params = {'type': "json",'buscar':palabra, 'proj': "MERC",'num':50,'key': "x4cGBaOC-423Z-A9ii-AKwM-QWiCfq63rMu9"}
            r = requests.post(url, params=params)
            resjson = r.json()
            for val in resjson['data']:
                string=val['ent_abr'].replace(".","")
                string=string.replace(" ","")
                if comparaestadosCosto(string)==datos['estado']:
                    data["cod"] = val['id_dest']
                    data["bandera"] = 2
                    return JsonResponse(data)

        data["bandera"] = 1
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/') 

def comparaestadosCosto(argument):
    switcher = {
        "Ags":"Aguascalientes",
        "BC":"Baja California",
        "BCS":"Baja California Sur",
        "Camp":"Campeche",
        "Coah":"Coahuila de Zaragoza",
        "Col":"Colima",
        "Chis":"Chiapas",
        "Chih":"Chihuahua",
        "CDMX":"Ciudad de México",
        "Dgo":"Durango",
        "Gto":"Guanajuato",
        "Gro":"Guerrero",
        "Hgo":"Hidalgo",
        "Jal":"Jalisco",
        "Mex":"México",
        "Mich":"Michoacán de Ocampo",
        "Mor":"Morelos",
        "Nay":"Nayarit",
        "NL":"Nuevo León",
        "Oax":"Oaxaca",
        "Pue":"Puebla",
        "Qro":"Querétaro",
        "QRoo":"Quintana Roo",
        "SLP":"San Luis Potosí",
        "Sin":"Sinaloa",
        "Son":"Sonora",
        "Tab":"Tabasco",
        "Tamps":"Tamaulipas",
        "Tlax":"Tlaxcala",
        "Ver":"Veracruz de Ignacio de la Llave",
        "Yuc":"Yucatán",
        "Zac":"Zacatecas",
    }
    return switcher.get(argument, "nothing")

#Ver costeo
def loadCosteo(request, id):
    if (request.user.is_authenticated):
        data = dict()
        costeo = get_object_or_404(Costeo, pk=id)
        data['Costeo_html'] = render_to_string('PartialCosteo.html', {'costeo': costeo})
        return JsonResponse(data)
    else:
        return HttpResponseRedirect('/') 

#Variables iniciales
def firstTime(request):
    if (request.user.is_authenticated):
        depreciacion = Depreciacion(
                    Meses = 60,
                    CostosUnidad = 800000.00,
                    CostosCaja = 90000.00,
                    ViajesMes=1,
                    KmsMesXunidad = 14000,
                    KmsMaximo = 15000,
                    DepTracto = .95,
                    DepCaja = .11,
                    RentaGPS = 400.00,
                    PlacasTenencia = 3000.00,
                    Seguro = 10000.00,
                    Admvo = 4,
                    Financieros = 2,
                    MttoUnidadXkm = 1.3,
                    Llantas = .3,
                    Operador = 1.3,
                    DobleOp = 1.3
                    )
        depreciacion.save()

        factoresPremisas = FactoresPremisas(
                    Unidad = 1,
                    Caja = 1,
                    KmSencillo = 1,
                    KmRoundTrip = 1,
                    KmMensuales = 1,
                    CasetaSingle = 1,
                    Rendimiento = 2.8,
                    Diesel = 1,
                    DieselSinIva = 1,
                    )
        factoresPremisas.save()

        
        costosOperativos = CostosOperativos(
                    Combustible = 7645.00,
                    Casetas = 3445.00,
                    Operador = 1520.00,
                    Subtotal1 = 12610.00,
                    MttoUnidad = 1520.00,
                    LlantasUnidad = 351.00,
                    Gps = 31.00,
                    Seguro = 779.00,
                    PlacasTenencia = 234.00,
                    SubTotal2 = 2915.00,
                    Admvo = 621.00,
                    Financieros = 310.00,
                    DeprUnidad = 1113.00,
                    DeprRemolque = 125.00,
                    Subtotal3 = 2170.00,
                    )
        costosOperativos.save()
    else:
        return HttpResponseRedirect('/') 